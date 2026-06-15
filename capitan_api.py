"""
CAPITAN AI — Enterprise Backend v29.0
CLOSEAI Technologies
FULL INTELLIGENCE RESTORED | Elite Reasoning | Human-Like Communication
All Rough Edges Fixed – File Analysis, Workspaces, Live Markets, Payments, Memories
"""

import os
import re
import json
import uuid
import time
import hmac
import hashlib
import base64
import secrets
import requests
import logging
import bcrypt
import PyPDF2
import docx
import openpyxl
import pandas as pd
import io
from datetime import datetime, timedelta, timezone
from typing import Optional, List, Tuple
from contextlib import contextmanager

from fastapi import FastAPI, HTTPException, Request, UploadFile, File, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, Response
from pydantic import BaseModel
from pydantic_settings import BaseSettings
import psycopg2
import uvicorn

# ================================================================
# FASTAPI APP
# ================================================================
app = FastAPI(title="CAPITAN AI API", version="29.0")

class Settings(BaseSettings):
    DATABASE_URL: str
    JWT_SECRET: str
    FOUNDER_KEY: str
    FRONTEND_URL: str = "https://capitanai.goldquantum0.workers.dev"
    GROQ_API_KEY: str = ""
    OPENROUTER_API_KEY: str = ""
    COINGECKO_KEY: str = ""
    SERPAPI_KEY: str = ""
    NEWS_API_KEY: str = ""
    FINNHUB_API_KEY: str = ""
    ETHERSCAN_API_KEY: str = ""

    class Config:
        env_file = ".env"
        extra = "ignore"

settings = Settings()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"]
)

# ================================================================
# LOGGING
# ================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ================================================================
# DATABASE – fixed context manager
# ================================================================
@contextmanager
def get_db():
    conn = None
    last_err = None
    for attempt in range(3):
        try:
            conn = psycopg2.connect(settings.DATABASE_URL, connect_timeout=10)
            break
        except Exception as e:
            last_err = e
            logger.warning(f"DB attempt {attempt+1} failed: {e}")
            if attempt < 2:
                time.sleep(2)
    if conn is None:
        raise last_err
    try:
        yield conn
    finally:
        conn.close()

def init_db():
    try:
        with get_db() as conn:
            with conn.cursor() as c:
                # ... (all existing tables unchanged) ...
                c.execute('''
                    CREATE TABLE IF NOT EXISTS users (
                        id UUID PRIMARY KEY,
                        email TEXT UNIQUE NOT NULL,
                        password_hash TEXT NOT NULL,
                        name TEXT,
                        tier TEXT DEFAULT 'free',
                        reasoning_depth INTEGER DEFAULT 1,
                        preferred_domain TEXT DEFAULT 'general',
                        daily_msg_count INTEGER DEFAULT 0,
                        msg_reset_date DATE,
                        tier_expires TIMESTAMP,
                        created_at TIMESTAMP DEFAULT NOW(),
                        updated_at TIMESTAMP DEFAULT NOW()
                    )
                ''')
                c.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS reasoning_depth INTEGER DEFAULT 1")
                c.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS preferred_domain TEXT DEFAULT 'general'")
                c.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS daily_msg_count INTEGER DEFAULT 0")
                c.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS msg_reset_date DATE")
                c.execute('''
                    CREATE TABLE IF NOT EXISTS user_sessions (
                        id UUID PRIMARY KEY,
                        user_id UUID REFERENCES users(id) ON DELETE CASCADE,
                        token TEXT UNIQUE NOT NULL,
                        expires_at TIMESTAMP,
                        created_at TIMESTAMP DEFAULT NOW()
                    )
                ''')
                c.execute('''
                    CREATE TABLE IF NOT EXISTS sessions (
                        id TEXT PRIMARY KEY,
                        tier TEXT DEFAULT 'free',
                        msg_count INTEGER DEFAULT 0,
                        daily_msg_count INTEGER DEFAULT 0,
                        msg_reset_date DATE,
                        created TIMESTAMP DEFAULT NOW(),
                        updated TIMESTAMP DEFAULT NOW()
                    )
                ''')
                c.execute("ALTER TABLE sessions ADD COLUMN IF NOT EXISTS daily_msg_count INTEGER DEFAULT 0")
                c.execute("ALTER TABLE sessions ADD COLUMN IF NOT EXISTS msg_reset_date DATE")
                c.execute('''
                    CREATE TABLE IF NOT EXISTS chats (
                        id TEXT PRIMARY KEY,
                        user_id UUID REFERENCES users(id) ON DELETE CASCADE,
                        session_id TEXT,
                        title TEXT,
                        created TIMESTAMP DEFAULT NOW(),
                        updated TIMESTAMP DEFAULT NOW()
                    )
                ''')
                c.execute('''
                    CREATE TABLE IF NOT EXISTS chat_messages (
                        id TEXT PRIMARY KEY,
                        chat_id TEXT,
                        user_id UUID REFERENCES users(id) ON DELETE CASCADE,
                        session_id TEXT,
                        role TEXT,
                        content TEXT,
                        model TEXT,
                        reasoning_chain TEXT,
                        created TIMESTAMP DEFAULT NOW()
                    )
                ''')
                c.execute("ALTER TABLE chat_messages ADD COLUMN IF NOT EXISTS reasoning_chain TEXT")
                c.execute('''
                    CREATE TABLE IF NOT EXISTS memories (
                        id TEXT PRIMARY KEY,
                        memory_id TEXT,
                        user_id UUID REFERENCES users(id) ON DELETE CASCADE,
                        content TEXT,
                        query TEXT,
                        domain TEXT,
                        importance INTEGER DEFAULT 1,
                        created TIMESTAMP DEFAULT NOW()
                    )
                ''')
                c.execute('''
                    CREATE TABLE IF NOT EXISTS library_items (
                        id TEXT PRIMARY KEY,
                        user_id UUID REFERENCES users(id) ON DELETE CASCADE,
                        name TEXT,
                        content TEXT,
                        created TIMESTAMP DEFAULT NOW()
                    )
                ''')
                c.execute('''
                    CREATE TABLE IF NOT EXISTS uploaded_files (
                        id TEXT PRIMARY KEY,
                        user_id UUID REFERENCES users(id) ON DELETE CASCADE,
                        filename TEXT,
                        original_name TEXT,
                        size INTEGER,
                        storage_path TEXT,
                        extracted_text TEXT,
                        created TIMESTAMP DEFAULT NOW()
                    )
                ''')
                c.execute("ALTER TABLE uploaded_files ADD COLUMN IF NOT EXISTS extracted_text TEXT")
                c.execute('''
                    CREATE TABLE IF NOT EXISTS workspaces (
                        id TEXT PRIMARY KEY,
                        name TEXT,
                        owner_id UUID REFERENCES users(id) ON DELETE CASCADE,
                        room_code TEXT UNIQUE,
                        max_members INTEGER DEFAULT 10,
                        created_at TIMESTAMP DEFAULT NOW()
                    )
                ''')
                c.execute('''
                    CREATE TABLE IF NOT EXISTS workspace_members (
                        workspace_id TEXT,
                        user_id UUID REFERENCES users(id) ON DELETE CASCADE,
                        role TEXT DEFAULT 'member',
                        joined_at TIMESTAMP DEFAULT NOW(),
                        PRIMARY KEY (workspace_id, user_id)
                    )
                ''')
                c.execute('''
                    CREATE TABLE IF NOT EXISTS workspace_messages (
                        id TEXT PRIMARY KEY,
                        workspace_id TEXT,
                        user_id UUID REFERENCES users(id) ON DELETE CASCADE,
                        author_name TEXT,
                        message TEXT,
                        is_ai INTEGER DEFAULT 0,
                        created TIMESTAMP DEFAULT NOW()
                    )
                ''')
                c.execute('''
                    CREATE TABLE IF NOT EXISTS payments (
                        id UUID PRIMARY KEY,
                        user_id UUID REFERENCES users(id) ON DELETE CASCADE,
                        txid TEXT UNIQUE,
                        currency TEXT,
                        amount REAL,
                        tier TEXT,
                        verified INTEGER DEFAULT 0,
                        created_at TIMESTAMP DEFAULT NOW()
                    )
                ''')
                c.execute('''
                    CREATE TABLE IF NOT EXISTS reasoning_cache (
                        id TEXT PRIMARY KEY,
                        query_hash TEXT UNIQUE,
                        reasoning_chain TEXT,
                        result TEXT,
                        created TIMESTAMP DEFAULT NOW()
                    )
                ''')
                conn.commit()
        logger.info("✅ Database ready")
    except Exception as e:
        logger.warning(f"DB init: {e}")

init_db()
def sid(): return str(uuid.uuid4())[:8].upper()
def mid(): return 'mem_' + sid()

# ... (all existing functions: hash_password, verify_password, JWT, etc. remain EXACTLY the same) ...

# ================================================================
# FILE EXTRACTION UTILITY
# ================================================================
def extract_text_from_file(file_path: str, original_name: str) -> str:
    """Extract text from uploaded files based on extension."""
    ext = original_name.rsplit('.', 1)[-1].lower() if '.' in original_name else ''
    try:
        if ext in ('txt', 'md', 'json', 'csv', 'py', 'js', 'html', 'css'):
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()
        elif ext == 'pdf':
            text = []
            with open(file_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    text.append(page.extract_text() or '')
            return '\n'.join(text)
        elif ext == 'docx':
            doc = docx.Document(file_path)
            return '\n'.join([p.text for p in doc.paragraphs])
        elif ext == 'xlsx':
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheets_text = []
            for name in wb.sheetnames:
                ws = wb[name]
                for row in ws.iter_rows(values_only=True):
                    row_text = ' '.join([str(c) if c is not None else '' for c in row])
                    sheets_text.append(row_text)
            return '\n'.join(sheets_text)
        else:
            return ''
    except Exception as e:
        logger.error(f"File extraction error: {e}")
        return ''

# ================================================================
# FILE UPLOAD (updated to extract text)
# ================================================================
@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    # ... (same checks) ...
    contents = await file.read()
    # ... (same size check, save to disk) ...
    extracted = extract_text_from_file(file_path, file.filename or "unknown")
    
    # Update DB with extracted text
    with get_db() as conn:
        with conn.cursor() as c:
            c.execute("UPDATE uploaded_files SET extracted_text = %s WHERE id = %s", (extracted[:50000], file_id))  # cap at 50k chars
            conn.commit()
    
    # The frontend will automatically send a user message with the file name; the backend will fetch the text in the chat endpoint.
    return {"id": file_id, "filename": file.filename, "size_mb": round(len(contents) / (1024 * 1024), 2), "extracted": bool(extracted)}

# ================================================================
# CHAT ENDPOINT (enhanced with file analysis & memories)
# ================================================================
@app.post("/api/chat")
async def chat_endpoint(req: ChatRequest, request: Request):
    # ... (same auth) ...
    user_msg = ... (get from messages)
    chat_id = req.chat_id or f"chat_{sid()}"
    
    # If user message mentions uploaded file, fetch the extracted text
    file_text = ""
    if "[Uploaded document:" in user_msg:
        # extract file name from message
        fname_match = re.search(r'\[Uploaded document:\s*(.*?)\]', user_msg)
        if fname_match:
            fname = fname_match.group(1).strip()
            with get_db() as conn:
                with conn.cursor() as c:
                    c.execute("SELECT extracted_text FROM uploaded_files WHERE original_name = %s AND user_id = %s ORDER BY created DESC LIMIT 1", (fname, user["id"]))
                    row = c.fetchone()
                    if row and row[0]:
                        file_text = row[0]
                        user_msg += "\n\n[DOCUMENT CONTENT]\n" + file_text[:30000]  # truncate for token limit
    
    # ... (save user message, fetch history, etc.) ...
    
    # Retrieve relevant memories
    memory_text = ""
    if is_authenticated:
        with get_db() as conn:
            with conn.cursor() as c:
                c.execute("SELECT content FROM memories WHERE user_id = %s AND domain = %s ORDER BY created DESC LIMIT 3", (user["id"], domain))
                rows = c.fetchall()
                if rows:
                    memory_text = "\n\n[RELEVANT MEMORIES]\n" + "\n".join([r[0][:200] for r in rows])
    
    # Build prompt with file content & memories
    prompt = build_system_prompt(domain, tier, tier_info["ai_model"], reasoning_depth, preferred_domain, web_results, user_query=user_msg)
    if memory_text:
        prompt += "\n" + memory_text
    # ... (call AI, save response) ...
    return {"content": result, ...}

# ================================================================
# MARKET DATA – replaced Yahoo with Finnhub
# ================================================================
def get_market_prices():
    results = {}
    # CoinGecko (crypto) – unchanged
    if settings.COINGECKO_KEY:
        # ... (same as before) ...
    
    # Finnhub for stocks/indices
    if settings.FINNHUB_API_KEY:
        symbols = ["AAPL", "MSFT", "NVDA", "TSLA", "GOOGL", "META", "AMZN", "^GSPC", "^IXIC", "^DJI"]
        for sym in symbols:
            try:
                r = requests.get(f"https://finnhub.io/api/v1/quote?symbol={sym}&token={settings.FINNHUB_API_KEY}", timeout=10)
                if r.status_code == 200:
                    data = r.json()
                    if data.get("c"):
                        results[sym] = {"price": data["c"], "change": round(data.get("dp", 0), 2)}
            except: pass
    return results

# ================================================================
# PAYMENT VERIFICATION – real blockchain check
# ================================================================
def verify_transaction(txid: str, currency: str, expected_tier: str) -> bool:
    prices = {"plus": 8, "pro": 17, "pro_max": 30}
    expected_amount = prices.get(expected_tier, 0)
    if currency == "BTC":
        try:
            r = requests.get(f"https://blockchain.info/rawtx/{txid}", timeout=15)
            if r.status_code == 200:
                tx = r.json()
                # check if any output matches our wallet and amount
                for out in tx.get("out", []):
                    if out.get("addr") == WALLETS["BTC"] and out.get("value", 0) / 1e8 >= expected_amount * 0.99:
                        return True
        except: pass
        return False
    elif currency == "ETH":
        if not settings.ETHERSCAN_API_KEY:
            return False
        try:
            r = requests.get(f"https://api.etherscan.io/api?module=proxy&action=eth_getTransactionByHash&txhash={txid}&apikey={settings.ETHERSCAN_API_KEY}", timeout=15)
            if r.status_code == 200:
                tx = r.json().get("result", {})
                if tx and tx.get("to", "").lower() == WALLETS["ETH"].lower():
                    value = int(tx.get("value", "0"), 16) / 1e18
                    if value >= expected_amount * 0.99:
                        return True
        except: pass
        return False
    return False

# ================================================================
# ADMIN – now with user management
# ================================================================
@app.get("/api/admin/users")
def admin_users(user: dict = Depends(get_current_user)):
    if not user or user["tier"] != "founder":
        raise HTTPException(403, "Access denied")
    with get_db() as conn:
        with conn.cursor() as c:
            c.execute("SELECT id, email, name, tier, created_at FROM users ORDER BY created_at DESC LIMIT 50")
            rows = c.fetchall()
            return [{"id": r[0], "email": r[1], "name": r[2], "tier": r[3], "created_at": r[4].isoformat() if r[4] else None} for r in rows]

@app.post("/api/admin/user/{user_id}/tier")
def admin_change_tier(user_id: str, req: dict, user: dict = Depends(get_current_user)):
    if not user or user["tier"] != "founder":
        raise HTTPException(403, "Access denied")
    new_tier = req.get("tier")
    if new_tier not in ("guest", "free", "plus", "pro", "pro_max", "founder"):
        raise HTTPException(400, "Invalid tier")
    with get_db() as conn:
        with conn.cursor() as c:
            c.execute("UPDATE users SET tier = %s, updated_at = NOW() WHERE id = %s", (new_tier, user_id))
            conn.commit()
    return {"success": True}

@app.delete("/api/admin/user/{user_id}")
def admin_delete_user(user_id: str, user: dict = Depends(get_current_user)):
    if not user or user["tier"] != "founder":
        raise HTTPException(403, "Access denied")
    with get_db() as conn:
        with conn.cursor() as c:
            c.execute("DELETE FROM users WHERE id = %s", (user_id,))
            conn.commit()
    return {"deleted": True}

# ================================================================
# RATE LIMITER CLEANUP
# ================================================================
rate_store = {}
_cleanup_counter = 0
def check_rate_limit(id: str, key: str = "default", limit: int = 20) -> bool:
    global _cleanup_counter
    now = time.time()
    store_key = f"rate:{key}:{id}"
    if store_key not in rate_store:
        rate_store[store_key] = []
    # clean expired entries every 100 requests
    _cleanup_counter += 1
    if _cleanup_counter % 100 == 0:
        for k in list(rate_store.keys()):
            rate_store[k] = [t for t in rate_store[k] if now - t < 120]  # 2 min window
            if not rate_store[k]:
                del rate_store[k]
    # limit check
    rate_store[store_key] = [t for t in rate_store[store_key] if now - t < 60]
    if len(rate_store[store_key]) >= limit:
        return False
    rate_store[store_key].append(now)
    return True

# ... (rest of the code remains identical) ...